from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class ParsedField:
    field_code: str
    field_name: str
    raw_value: str
    normalized_value: str
    unit: str | None
    source_path: str
    confidence: int


class EtabsParser:
    """Extract a controlled summary from ETABS Excel exports."""

    FIELD_CODES = {
        "sheet_count": "ETABS-SHEET-COUNT",
        "table_names": "ETABS-TABLE-NAMES",
        "story_count": "ETABS-STORY-COUNT",
        "frame_object_count": "ETABS-FRAME-OBJECT-COUNT",
        "point_object_count": "ETABS-POINT-OBJECT-COUNT",
        "material_table_count": "ETABS-MATERIAL-TABLE-COUNT",
        "section_table_count": "ETABS-SECTION-TABLE-COUNT",
    }

    def parse(self, file_path: str) -> dict:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(file_path)
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise ValueError("ETABS parser expects an Excel export")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        fields = [
            self._field("sheet_count", len(sheet_names), str(path), "count", 95),
            self._field("table_names", ", ".join(sheet_names), str(path), None, 90),
            self._field("story_count", self._count_rows_matching_sheet(workbook, "story"), str(path), "count", 80),
            self._field("frame_object_count", self._count_rows_matching_sheet(workbook, "frame"), str(path), "count", 80),
            self._field("point_object_count", self._count_rows_matching_sheet(workbook, "point"), str(path), "count", 80),
            self._field("material_table_count", self._count_sheets_containing(sheet_names, "material"), str(path), "count", 85),
            self._field("section_table_count", self._count_sheets_containing(sheet_names, "section"), str(path), "count", 85),
        ]
        workbook.close()

        fields = [field for field in fields if field.raw_value not in ("0", "")]
        return {
            "parser": "EtabsParser",
            "source_path": str(path),
            "field_count": len(fields),
            "confidence": 85 if fields else 0,
            "fields": [field.__dict__ for field in fields],
        }

    def _field(self, field_name: str, value: object, source_path: str, unit: str | None, confidence: int) -> ParsedField:
        return ParsedField(
            field_code=self.FIELD_CODES[field_name],
            field_name=field_name,
            raw_value=str(value),
            normalized_value=str(value),
            unit=unit,
            source_path=source_path,
            confidence=confidence,
        )

    def _count_sheets_containing(self, sheet_names: list[str], token: str) -> int:
        return sum(1 for name in sheet_names if token in name.lower())

    def _count_rows_matching_sheet(self, workbook, token: str) -> int:
        matching_sheets = [name for name in workbook.sheetnames if token in name.lower()]
        if not matching_sheets:
            return 0
        count = 0
        for sheet_name in matching_sheets:
            worksheet = workbook[sheet_name]
            count += max((worksheet.max_row or 1) - 1, 0)
        return count
